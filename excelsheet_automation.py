import openpyxl

inv_file= openpyxl.load_workbook('inventory.xlsx')

product_list= inv_file['Sheet1']

products_per_supplier= {}

print(product_list.max_row)

#List each company with respective number of products

for product_row in range(2,product_list.max_row+1 ):
         supplier_name=product_list.cell(product_row,4).value

         if supplier_name in products_per_supplier:
                current_num_products= products_per_supplier[supplier_name] 
                products_per_supplier[supplier_name] = products_per_supplier[supplier_name] +1
         else:    
                print("adding new supplier to dictionary")
                products_per_supplier[supplier_name]=1




# calculate total inventory value per supplier

total_value_per_supplier= {}

for product_row in range(2,product_list.max_row+1 ):
              supplier_name = product_list.cell(product_row,4).value
              inventory= product_list.cell(product_row,2).value
              price= product_list.cell(product_row,3).value

              if supplier_name in total_value_per_supplier:
                     current_total_value= total_value_per_supplier[supplier_name]
                     total_value_per_supplier[supplier_name]= current_total_value + (inventory * price)
              else:
                     total_value_per_supplier[supplier_name]= inventory * price




# Listing all products with inventory less than 10

product_under_10_inv= {}

for product_row in range(2,product_list.max_row+1 ):
              product_name= product_list.cell(product_row,4).value
              inventory= product_list.cell(product_row,2).value
              product_number = product_list.cell(product_row,1).value
              if inventory < 10:
                     product_under_10_inv[int(product_number)]= int(inventory)



# Adding a new column to the excel sheet

for product_row in range(2,product_list.max_row+1 ):
              product_name= product_list.cell(product_row,4).value
              inventory= product_list.cell(product_row,2).value
              product_number = product_list.cell(product_row,1).value
              inventory_price = product_list.cell(product_row,5)
              inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)
inv_file.save('inventory_with_total_value.xlsx')
